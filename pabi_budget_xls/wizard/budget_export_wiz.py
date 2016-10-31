# -*- coding: utf-8 -*-
import base64
import cStringIO

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname
from openpyxl.styles import PatternFill, Border,\
    Side, Protection, Font

from openerp import models, fields, api, _
from openerp.exceptions import Warning as UserError


class BudgetExportWizard(models.Model):
    _name = 'unit.budget.plan.export'

    attachment_id = fields.Many2one(
        'ir.attachment',
        string='Template Plan',
    )
    editable_lines = fields.Integer(
        string='Additional Budget lines',
        required=True,
        default=10,
    )

    @api.model
    def _add_cell_border(self, sheet, row_start, row_end, col_start, col_end):
        if not row_start:
            row_start = 1
        if not col_start:
            col_start = 0
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        row = row_start
        for row in range(row_start, row_end):
            for col in range(col_start, col_end+1):
                sheet.cell(row=row, column=col).border = border

    @api.model
    def _make_cell_editable(self, sheet, row_start, row_end,
                            col_start, col_end, skip_cell):
        if not row_start:
            row_start = 1
        if not col_start:
            col_start = 0
        protection = Protection(locked=False)
        row = row_start
        for row in range(row_start, row_end):
            for col in range(col_start, col_end+1):
                if skip_cell and col == skip_cell:
                    continue
                sheet.cell(row=row, column=col).protection = protection

    @api.model
    def _make_cell_color_filled(self, sheet, row_start, row_end,
                                col_start, col_end, col_list):
        greyFill = PatternFill(
            start_color='D3D3D3',
            end_color='D3D3D3',
            fill_type='solid',
        )
        row = row_start
        if not col_list:
            col_list = range(col_start, col_end+1)

        for row in range(row_start, row_end):
            for col in col_list:
                sheet.cell(row=row, column=col).fill = greyFill

    @api.multi
    def update_budget_xls(self, budget_ids, template_id=None):
        if not template_id:
            raise UserError(_('Please add .xlsx template.'))
        if not budget_ids:
            raise UserError(_('No budget to export.'))
        budgets = self.env['budget.plan.unit'].browse(budget_ids)
        template_file = self.env['ir.attachment'].browse(template_id)

        for budget in budgets:
            exp_file = self.attachment_id.datas.decode('base64')
            stream = cStringIO.StringIO(exp_file)
            workbook = openpyxl.load_workbook(stream)

            AG_Sheet = workbook.get_sheet_by_name('ActivityGroup_MasterData')
            AG_Sheet.protection.sheet = True
            activities = self.env['account.activity.group'].search([])

            bold_font = Font(bold=True, name='Arial', size=11)

            AG_Sheet.cell(row=1, column=1).value = 'Sequence'
            AG_Sheet.cell(row=1, column=2).value = 'Activity Group - English'
            AG_Sheet.cell(row=1, column=3).value = 'Activity Group - Thai'
            AG_Sheet.cell(row=1, column=1).font = bold_font
            AG_Sheet.cell(row=1, column=2).font = bold_font
            AG_Sheet.cell(row=1, column=3).font = bold_font

            ag_row = 2
            ag_count = 1
            ag_length = 1
            for ag in activities:
                AG_Sheet.cell(row=ag_row, column=1, value=ag_count)
                AG_Sheet.cell(row=ag_row, column=2, value=ag.name)
                AG_Sheet.cell(row=ag_row, column=3, value=ag.name)
                if len(ag.name) > ag_length:
                    ag_length = len(ag.name)
                ag_row += 1
                ag_count += 1

            AG_Sheet.column_dimensions['A'].width = 11
            AG_Sheet.column_dimensions['B'].width = ag_length
            AG_Sheet.column_dimensions['C'].width = ag_length

            formula1 = "{0}!$C$2:$C$%s" % (ag_row)
            dv = DataValidation(
                type="list",
                formula1=formula1.format(
                    quote_sheetname('ActivityGroup_MasterData')
                )
            )
            NonCostCtrl_Sheet =\
                workbook.get_sheet_by_name('Non_CostControl')
            NonCostCtrl_Sheet.protection.sheet = True

            NonCostCtrl_Sheet.add_data_validation(dv)
            NonCostCtrl_Sheet.cell(row=1, column=5, value=budget.id)
            org = budget.org_id.code and\
                    budget.org_id.code or budget.org_id.name_short
            NonCostCtrl_Sheet.cell(row=1, column=2,
                                   value=budget.fiscalyear_id.name)
            NonCostCtrl_Sheet.cell(row=2, column=2, value=org)
            NonCostCtrl_Sheet.cell(row=3, column=2,
                                   value=budget.section_id.code)
            NonCostCtrl_Sheet.cell(row=4, column=2, value=fields.Date.today())
            NonCostCtrl_Sheet.cell(row=5, column=2, value=self.env.user.name)
            NonCostCtrl_Sheet.cell(row=6, column=2,
                                   value=str(budget.planned_overall))

            row = 11
            section_name = budget.section_id.name
            LineStart = row
            for line in budget.plan_line_ids:
                if line.section_id:
                    section_name = line.section_id.name
                NonCostCtrl_Sheet.cell(row=row, column=1).value = '=B2'
                NonCostCtrl_Sheet.cell(row=row, column=2).value = "=B3"
                NonCostCtrl_Sheet.cell(
                    row=row, column=3).value = line.section_id.name
                NonCostCtrl_Sheet.cell(
                    row=row, column=7).value = line.unit
                NonCostCtrl_Sheet.cell(
                    row=row, column=8).value = line.activity_unit_price
                NonCostCtrl_Sheet.cell(
                    row=row, column=9).value = line.activity_unit
                dv.add(NonCostCtrl_Sheet.cell(row=row, column=4))
                ag_name = line.activity_group_id.name
                NonCostCtrl_Sheet.cell(row=row, column=4).value = ag_name
                NonCostCtrl_Sheet.cell(
                    row=row, column=10).value = "=G%s*$H$%s*$I$%s" % (row,
                                                                      row,
                                                                      row)

                NonCostCtrl_Sheet.cell(row=row, column=12).value = line.m1
                NonCostCtrl_Sheet.cell(row=row, column=13).value = line.m2
                NonCostCtrl_Sheet.cell(row=row, column=14).value = line.m3
                NonCostCtrl_Sheet.cell(row=row, column=15).value = line.m4
                NonCostCtrl_Sheet.cell(row=row, column=16).value = line.m5
                NonCostCtrl_Sheet.cell(row=row, column=17).value = line.m6
                NonCostCtrl_Sheet.cell(row=row, column=18).value = line.m7
                NonCostCtrl_Sheet.cell(row=row, column=19).value = line.m8
                NonCostCtrl_Sheet.cell(row=row, column=20).value = line.m9
                NonCostCtrl_Sheet.cell(row=row, column=21).value = line.m10
                NonCostCtrl_Sheet.cell(row=row, column=22).value = line.m11
                NonCostCtrl_Sheet.cell(row=row, column=23).value = line.m12
                NonCostCtrl_Sheet.cell(
                    row=row, column=24, value="=SUM(L%s:$W$%s)" % (row, row))
                NonCostCtrl_Sheet.cell(
                    row=row, column=25, value="=J%s-$X$%s" % (row, row))
                NonCostCtrl_Sheet.cell(row=row, column=26).value = line.id
                row += 1

            to_row = row + self.editable_lines
            for r in range(row, to_row):
                NonCostCtrl_Sheet.cell(row=r, column=1).value = "=B2"
                NonCostCtrl_Sheet.cell(row=r, column=2).value = "=B3"
                NonCostCtrl_Sheet.cell(
                    row=r, column=3).value = section_name
                dv.add(NonCostCtrl_Sheet.cell(row=r, column=4))
                NonCostCtrl_Sheet.cell(
                    row=r, column=10).value = "=G%s*$H$%s*$I$%s" % (r, r, r)
                NonCostCtrl_Sheet.cell(
                    row=r, column=24, value="=SUM(L%s:$W$%s)" % (r, r))
                NonCostCtrl_Sheet.cell(
                    row=r, column=25, value="=J%s-$X$%s" % (r, r))
                row += 1
                r += 1

            column_to_fill = [1, 2, 3, 10, 11, 24, 25]
            self._add_cell_border(NonCostCtrl_Sheet,
                                  row_start=LineStart,
                                  row_end=row,
                                  col_start=1,
                                  col_end=25)
            self._make_cell_editable(sheet=NonCostCtrl_Sheet,
                                     row_start=LineStart,
                                     row_end=row,
                                     col_start=4,
                                     col_end=23,
                                     skip_cell=10)
            self._make_cell_color_filled(sheet=NonCostCtrl_Sheet,
                                         row_start=LineStart,
                                         row_end=row,
                                         col_start=1,
                                         col_end=3,
                                         col_list=column_to_fill)

            NonCostCtrl_Sheet.cell(row=row, column=9).value = 'Total'
            NonCostCtrl_Sheet.cell(row=row, column=9).font = bold_font
            params = (LineStart, row-1)
            NonCostCtrl_Sheet.cell(
                row=row, column=10).value = '=SUM(J%s:J%s)' % params
            NonCostCtrl_Sheet.cell(
                row=row, column=11).value = '=SUM(K%s:K%s)' % params
            NonCostCtrl_Sheet.cell(
                row=row, column=12).value = '=SUM(L%s:L%s)' % params
            NonCostCtrl_Sheet.cell(
                row=row, column=13).value = '=SUM(M%s:M%s)' % params
            NonCostCtrl_Sheet.cell(
                row=row, column=14).value = '=SUM(N%s:N%s)' % params
            NonCostCtrl_Sheet.cell(
                row=row, column=15).value = '=SUM(O%s:O%s)' % params
            NonCostCtrl_Sheet.cell(
                row=row, column=16).value = '=SUM(P%s:P%s)' % params
            NonCostCtrl_Sheet.cell(
                row=row, column=17).value = '=SUM(Q%s:Q%s)' % params
            NonCostCtrl_Sheet.cell(
                row=row, column=18).value = '=SUM(R%s:R%s)' % params
            NonCostCtrl_Sheet.cell(
                row=row, column=19).value = '=SUM(S%s:S%s)' % params
            NonCostCtrl_Sheet.cell(
                row=row, column=20).value = '=SUM(T%s:T%s)' % params
            NonCostCtrl_Sheet.cell(
                row=row, column=21).value = '=SUM(U%s:U%s)' % params
            NonCostCtrl_Sheet.cell(
                row=row, column=22).value = '=SUM(V%s:V%s)' % params
            NonCostCtrl_Sheet.cell(
                row=row, column=23).value = '=SUM(W%s:W%s)' % params

            self._add_cell_border(NonCostCtrl_Sheet, row_start=row,
                                  row_end=row+1, col_start=9, col_end=23)
            self._make_cell_color_filled(sheet=NonCostCtrl_Sheet,
                                         row_start=row, row_end=row+1,
                                         col_start=9, col_end=23, col_list=[])

            stream1 = cStringIO.StringIO()
            workbook.save(stream1)
            filename = '%s-%s-%s-%s.xlsx' % (budget.fiscalyear_id.name,
                                             org,
                                             budget.section_id.code,
                                             template_file.name)
            self.env.cr.execute(""" DELETE FROM budget_xls_output """)
            attachement_id = self.env['ir.attachment'].create({
                'name': filename,
                'datas': stream1.getvalue().encode('base64'),
                'datas_fname': filename,
                'res_model': 'budget.plan.unit',
                'res_id': budget.id,
            })
            attach_id = self.env['budget.xls.output'].create(
                {'name': filename,
                 'xls_output': base64.encodestring(stream1.getvalue()),
                 }
            )
            self.env['budget.plan.history'].create({
                'user_id': self.env.user.id,
                'operation_date': fields.Datetime.now(),
                'operation_type': 'export',
                'plan_id': budget.id,
                'attachement_id': attachement_id.id
            })
            return attach_id

    @api.multi
    def export_budget(self):
        budget_ids = self.env.context.get('active_ids')
        attach_id = self.update_budget_xls(budget_ids, self.attachment_id.id)
        return {
            'context': self.env.context,
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'budget.xls.output',
            'res_id': attach_id.id,
            'type': 'ir.actions.act_window',
            'target': 'new',
        }
